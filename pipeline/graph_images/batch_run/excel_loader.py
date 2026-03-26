from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

from common import RowData, normalize_header


HEADER_ALIASES = {
    "sku": {"sku code", "sku", "sku id"},
    "product_name": {"product name", "name"},
    "category": {"category"},
    "hero_image": {"hero image", "hero image link", "hero image url"},
    "motif_width": {"motif width (mm)", "motif width mm", "motif width"},
    "overall_width": {"overall width(mm)", "overall width (mm)", "overall width mm", "overall width"},
    "overall_height": {"overall height (mm)", "overall height(mm)", "overall height mm", "overall height"},
    "chain_length": {"chain (cm)", "chain length", "chain length (cm)", "length of chain", "chain"},
    "adjustable": {"adjustable", "adjustable or not", "is adjustable"},
}


def _find_column_indexes(sheet) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for col_idx in range(1, sheet.max_column + 1):
        normalized = normalize_header(sheet.cell(1, col_idx).value)
        for canonical, aliases in HEADER_ALIASES.items():
            if normalized in aliases:
                header_map[canonical] = col_idx
    return header_map


def _cell_numeric(cell_value: Any) -> float | None:
    if cell_value in (None, ""):
        return None
    try:
        return float(cell_value)
    except (TypeError, ValueError):
        return None


def _cell_text(cell_value: Any) -> str:
    if cell_value is None:
        return ""
    return str(cell_value).strip()


def _hero_image_url(cell) -> str:
    hyperlink = getattr(cell, "hyperlink", None)
    if hyperlink is not None and hyperlink.target:
        return str(hyperlink.target).strip()
    return _cell_text(cell.value)


def load_rows(excel_path: Path) -> list[RowData]:
    workbook = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    indexes = _find_column_indexes(sheet)

    rows: list[RowData] = []
    for row_idx in range(2, sheet.max_row + 1):
        sku = _cell_text(sheet.cell(row_idx, indexes.get("sku", 0)).value) if indexes.get("sku") else ""
        product_name = _cell_text(sheet.cell(row_idx, indexes.get("product_name", 0)).value) if indexes.get("product_name") else ""
        category = _cell_text(sheet.cell(row_idx, indexes.get("category", 0)).value) if indexes.get("category") else ""

        if not any([sku, product_name, category]):
            continue

        hero_cell = sheet.cell(row_idx, indexes.get("hero_image", 0)) if indexes.get("hero_image") else None
        hero_image_url = _hero_image_url(hero_cell) if hero_cell is not None else ""

        raw: dict[str, Any] = {}
        for key, col_idx in indexes.items():
            raw[key] = sheet.cell(row_idx, col_idx).value

        rows.append(
            RowData(
                row_index=row_idx,
                sku=sku or f"row_{row_idx}",
                product_name=product_name or f"Row {row_idx}",
                category=category,
                hero_image_url=hero_image_url,
                motif_width_mm=_cell_numeric(raw.get("motif_width")),
                overall_width_mm=_cell_numeric(raw.get("overall_width")),
                overall_height_mm=_cell_numeric(raw.get("overall_height")),
                chain_length_cm=_cell_numeric(raw.get("chain_length")),
                adjustable=str(raw.get("adjustable", "")).strip().lower() in {"yes", "y", "true", "1", "adjustable"},
                raw=raw,
            )
        )
    return rows

